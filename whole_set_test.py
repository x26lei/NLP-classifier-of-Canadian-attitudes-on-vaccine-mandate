from openpyxl.reader.excel import load_workbook
from random_batch import *
import numpy as np
import torch
import random


def load_data(excel_yes, excel_no):
    """读取数据"""
    wb_yes = load_workbook(excel_yes)
    ws_yes = wb_yes.active
    wb_no = load_workbook(excel_no)
    ws_no = wb_no.active
    """读取数据"""
    list_yes = []
    for i in range(1, 2555):
        embedding_yes = ws_yes.cell(row=i, column=2).value
        embedding_yes = embedding_yes.lstrip('[')
        embedding_yes = embedding_yes.rstrip(']')
        em = str.split(embedding_yes)
        for j in range(len(em)):
            em[j] = eval(em[j])
        em = np.array(em, dtype='float32')
        em = torch.from_numpy(em)
        list_yes.append(em)
    list_no = []
    for i in range(1, 11369):
        embedding_no = ws_no.cell(row=i, column=2).value
        embedding_no = embedding_no.lstrip('[')
        embedding_no = embedding_no.rstrip(']')
        em = str.split(embedding_no)
        for j in range(len(em)):
            em[j] = eval(em[j])
        em = np.array(em, dtype='float32')
        em = torch.from_numpy(em)
        list_no.append(em)
    x_t = list_yes + list_no
    y_t = [1] * 2554 + [0] * 11368
    return x_t, y_t


net1 = torch.load('net3_26_1.pkl')
net1.eval()
x_test, y_test= load_data('Yes_Sep.xlsx', 'No_Sep.xlsx')
test_data, test_label = random_batch(x_test, y_test, len(x_test))
y_p = net1(test_data[0])
y_h = test_label[0]
summ = 0
y_p_l = torch.zeros((y_h.shape[0]))
for i in range (0, y_h.shape[0]):
    if y_p[i,0] >= 0.5:
        y_p_l[i] = 0
    else:
        y_p_l[i] = 1
    if y_p_l[i] == y_h[i]:
        summ = summ + 1
acc = summ / y_h.shape[0]
print(acc)

wb = load_workbook('Yes_Sep.xlsx')
ws = wb.active
for i in range (1,2555):
    ws.cell(i,3,value = str(y_p_l[i-1].item()))
wb.save('Yes_Sep.xlsx')

wb = load_workbook('No_Sep.xlsx')
ws = wb.active
for i in range (1,11369):
    ws.cell(i,3,value =  str(y_p_l[i - 1 + 2554].item()))
wb.save('No_Sep.xlsx')