from openpyxl.reader.excel import load_workbook
import numpy as np
import torch
import random


def load_data(excel_yes, excel_no, number_training_examples):
    """读取数据"""
    wb_yes = load_workbook(excel_yes)
    ws_yes = wb_yes.active
    wb_no = load_workbook(excel_no)
    ws_no = wb_no.active
    """读取数据"""
    list_yes = []
    for i in range(1, 2555):
        embedding_yes = ws_yes.cell(row=i, column=2).value
        embedding_yes = embedding_yes.lstrip('[')
        embedding_yes = embedding_yes.rstrip(']')
        em = str.split(embedding_yes)
        for j in range(len(em)):
            em[j] = eval(em[j])
        em = np.array(em, dtype='float32')
        em = torch.from_numpy(em)
        list_yes.append(em)
    list_no = []
    # for i in range(1, 11369):
    for i in range(1, 11369):
        embedding_no = ws_no.cell(row=i, column=2).value
        embedding_no = embedding_no.lstrip('[')
        embedding_no = embedding_no.rstrip(']')
        em = str.split(embedding_no)
        for j in range(len(em)):
            em[j] = eval(em[j])
        em = np.array(em, dtype='float32')
        em = torch.from_numpy(em)
        list_no.append(em)
    x_whole = list_yes + list_no
    y_whole = [1] * 2554 + [0] * 11368
    x_t = list_yes + list_no[0: 2554]
    y_t = [1] * 2554 + [0] * 2554
    randnum = random.randint(0, 100)
    random.seed(randnum)
    random.shuffle(x_t)
    random.seed(randnum)
    random.shuffle(y_t)
    x_train = x_t[0: number_training_examples]
    y_train = y_t[0: number_training_examples]
    x_test = x_t[number_training_examples:]
    y_test = y_t[number_training_examples:]
    # print(len(x_train), len(y_train), len(x_test), len(y_test))
    return x_train, y_train, x_test, y_test, x_whole, y_whole
