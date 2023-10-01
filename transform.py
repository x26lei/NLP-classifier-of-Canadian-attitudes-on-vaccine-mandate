from sentence_transformers import SentenceTransformer
from openpyxl.reader.excel import load_workbook
'''
model = SentenceTransformer('sentence-transformers/all-distilroberta-v1')
wb = load_workbook('Yes_Sep.xlsxYes_Sep.xlsx')
ws = wb.active
for i in range (1,2555):
    s = ws.cell(row = i, column = 1).value
    embeddings = model.encode(s)
    ws.cell(i,2,value = str(embeddings))
wb.save('Yes_Sep.xlsx')

wb = load_workbook('No_Sep.xlsx')
ws = wb.active
for i in range (1,11369):
    s = ws.cell(row = i, column = 1).value
    embeddings = model.encode(s)
    ws.cell(i,2,value = str(embeddings))
wb.save('No_Sep.xlsx')
'''
model = SentenceTransformer('sentence-transformers/all-distilroberta-v1')
wb = load_workbook('data/britishcolumbia,_KW_Comments_October.xlsx')
ws = wb.active
#for i in range (2,848):
for i in range (1,10):
    print(i)
    s = ws.cell(row = i, column = 3).value
    embeddings = model.encode(s)
    ws.cell(i,17,value = str(embeddings))
wb.save('data/britishcolumbia,_KW_Comments_October.xlsx')
