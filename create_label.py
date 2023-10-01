from openpyxl.reader.excel import load_workbook
import numpy as np
import torch
import random


def load_data(excel_yes):
    """读取数据"""
    wb_yes = load_workbook(excel_yes)
    ws_yes = wb_yes.active
    """读取数据"""
    list_yes = []
    for i in range(2, 844):
        embedding_yes = ws_yes.cell(row=i, column=17).value
        embedding_yes = embedding_yes.lstrip('[')
        embedding_yes = embedding_yes.rstrip(']')
        em = str.split(embedding_yes)
        for j in range(len(em)):
            em[j] = eval(em[j])
        em = np.array(em, dtype='float32')
        em = torch.from_numpy(em)
        list_yes.append(em)
        x_t = list_yes
    return x_t

def random_batch(x_t, batch_size):
    '''
    randnum = random.randint(0, 100)
    random.seed(randnum)
    random.shuffle(x_t)
    random.seed(randnum)
    random.shuffle(y_t)
    '''
    batch = []
    for i in range(0, int(len(x_t) / batch_size)):
        left = 0 + i * batch_size
        right = batch_size + i * batch_size
        cut_x = x_t[left:right]
        batch_i = torch.zeros([batch_size, 768])
        for j in range(0, len(batch_i)):
            cut_x[j] = cut_x[j][None, :]
            batch_i[j, :] = cut_x[j]
        batch.append(batch_i)
    return batch


net1 = torch.load('net3_26_1.pkl')
net1.eval()
x_t= load_data('data/britishcolumbia,_KW_Comments_October.xlsx')
test_data, = random_batch(x_t, len(x_t))
print(test_data)
y_p = net1(test_data)
summ = 0
y_p_l = torch.zeros((y_p.shape[0]))
for i in range (0, y_p.shape[0]):
    if y_p[i,0] >= 0.5:
        y_p_l[i] = 0
    else:
        y_p_l[i] = 1
print(y_p_l.shape)
wb = load_workbook('data/britishcolumbia,_KW_Comments_October.xlsx')
ws = wb.active
for i in range (2,844):
    ws.cell(i,18,value = str(y_p_l[i-2].item()))
wb.save('data/britishcolumbia,_KW_Comments_October.xlsx')